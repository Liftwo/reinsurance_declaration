import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

df = pd.read_excel('D:\\團傷臨分明細.xlsx', '出單')
df = df.fillna('n')
dfX = (df.loc[(df['合約年'] == 2019)])

df2 = dfX[['分保合約編號', '再保人', '合約名稱','保單號碼','再保費', '保險金額', '保險期間']]
df2.rename(columns={'分保合約編號':'臨分合約編號', '合約名稱':'被保人', '再保費':'分出再保費(A)','保險金額':'再保分出限額(B)'}, inplace=True)

df2 = df2[df2['保單號碼'] != '合約']
df2['分出再保費(A)'] = df2.apply(lambda x: '{:,.0f}'.format(x['分出再保費(A)']), axis=1)

df2.insert(1, '合約類別', '比例再保險')
rank = {'starr': 'AM Best-A', 'Starr':'AM BEST-A','marsh':'broker', '聯聿':'broker', '華南':'AM Best-A'}
rank_col = []
reinsurer = {'starr': 'Starr International Insurance(Asia) Ltd.',
             'Starr':'Starr International Insurance(Asia) Ltd.', 'marsh':'Marsh Ltd.Taiwan Branch',
             '聯聿':'Lian Yu Insurance Brokers Co., Ltd.', '華南':'華南產險股份有限公司'}
reinsurer_col = []
for i in df2['再保人']:
    if i in rank:
        rank_col.append(rank[i])
        reinsurer_col.append(reinsurer[i])
    else:
        rank_col.append('n')
        reinsurer_col.append('n')

df2 = df2.drop('再保人', axis=1)
df2.insert(2, '再保人', reinsurer_col)
df2.insert(3, '信用評等', rank_col)
df2.insert(6, '再保險人實質上已承擔與原保險契約再保分出部分相關之所有保險危險', 'V')
df2.insert(10, '(A)/(B)', 0)
df2.insert(11, '是否通過', '是')

df2.insert(12, '生效日', df2['保險期間'].str.split('-').str[0])
df2['生效日'] = pd.to_datetime(df2['生效日'], errors='coerce')
df2['測試時間'] = df2['生效日'] + pd.DateOffset(days=-1)

df2['生效日'] = df2['生效日'].dt.strftime('%Y/%m/%d')
df2['測試時間'] = df2['測試時間'].dt.strftime('%Y/%m/%d')

df2.insert(13, '製表日期',  df2['測試時間'])
df2 = df2.drop('保險期間', axis=1)
df3 = pd.DataFrame(columns=['remark'], data=['註:合約類別指比例性再保或非比例性再保'])
df4 = pd.concat([df2, df3.rename(columns={'remark':'臨分合約編號'})], ignore_index=True, sort=False)

writer = pd.ExcelWriter('D:\\風險測試表1.xlsx', engine='openpyxl')
df4.to_excel(writer, '2019年分出臨分簡易測試表', startrow=2, index=False)
writer.save()


wb = load_workbook('D:\\風險測試表1.xlsx')
ws = wb.active

for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=6):
    for cell in rows:
        cell.fill = PatternFill(fgColor='FF99CE', fill_type='solid')
for rows in ws.iter_rows(min_row=3, max_row=3, min_col=11, max_col=14):
    for cell in rows:
        cell.fill = PatternFill(fgColor='CC99FF', fill_type='solid')
ws['G3'].fill = PatternFill(fgColor='FFFF00', fill_type='solid')
Orangefill = PatternFill(fgColor='FFC000', fill_type='solid')
ws['H3'].fill = Orangefill
ws['I3'].fill = Orangefill
ws['J3'].fill = Orangefill

border = Border(left=Side(border_style='thin', color='000000'),
         right=Side(border_style='thin', color='000000'),
         top=Side(border_style='thin', color='000000'),
         bottom=Side(border_style='thin', color='000000'))
for row in ws[3:ws.max_row-1]:
    for cell in row:
        cell.border = border
ws['A1'] = '分出臨分簡易測試表'
ws['A1'].font = Font(bold=True)
ws['A2'] = '2019'
ws['A2'].alignment = Alignment(horizontal='left')
ws['A2'].font = Font(bold=True, underline='single')

wb.save('D:\風險測試表2.xlsx')
wb.close()















